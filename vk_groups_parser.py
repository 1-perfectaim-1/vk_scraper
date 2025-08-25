#!/usr/bin/env python3
import argparse
import csv
import os
import sys
import time
import urllib.parse
import re
import math
from dataclasses import dataclass
from typing import Iterable, List, Optional, Set, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

from tqdm import tqdm

# Excel
from openpyxl import Workbook, load_workbook

# Playwright
from playwright.sync_api import Playwright, sync_playwright, BrowserContext, Page, TimeoutError as PlaywrightTimeoutError
from urllib.parse import urlparse, urlunparse


SEARCH_URL_TEMPLATE = "https://vk.com/search?c%5Bq%5D={query}&c%5Bsection%5D=communities"
VK_GROUPS_URL = "https://vk.com/groups"
STORAGE_STATE_FILE = "./.vk_storage_state.json"


@dataclass
class GroupItem:
    index: int
    name: str
    url: str
    subscribers: Optional[int]
    posting_status: Optional[str]


csv_lock = threading.Lock()
index_lock = threading.Lock()
seen_lock = threading.Lock()


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def _read_csv_header(file_path: str) -> Optional[List[str]]:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            return next(reader)
    except FileNotFoundError:
        return None
    except StopIteration:
        return []


def ensure_csv(file_path: str) -> None:
    ensure_parent_dir(file_path)
    desired = ["№", "Название", "Ссылка", "Подписчики", "Публикация"]
    header = _read_csv_header(file_path)
    if header is None:
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(desired)
    else:
        if not header:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(desired)
        else:
            missing = [c for c in desired if c not in header]
            if missing:
                # Upgrade: add missing columns and pad rows
                rows: List[List[str]] = []
                with open(file_path, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                # Build new header preserving order of existing and appending missing in desired order
                new_header = header[:]
                for c in desired:
                    if c not in new_header:
                        new_header.append(c)
                # Map old -> new width
                for i in range(1, len(rows)):
                    while len(rows[i]) < len(new_header):
                        rows[i].append("")
                rows[0] = new_header
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(rows)


def ensure_xlsx(file_path: str) -> None:
    ensure_parent_dir(file_path)
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "groups"
        ws.append(["№", "Название", "Ссылка", "Подписчики", "Публикация"])
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        changed = False
        for col in ["Подписчики", "Публикация"]:
            if col not in headers:
                headers.append(col)
                changed = True
        if changed:
            # rewrite header row
            for c, val in enumerate(headers, start=1):
                ws.cell(row=1, column=c, value=val)
            wb.save(file_path)


def append_csv(file_path: str, item: GroupItem) -> None:
    with csv_lock:
        with open(file_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                item.index,
                item.name,
                normalize_vk_url(item.url),
                item.subscribers if item.subscribers is not None else "",
                item.posting_status if item.posting_status is not None else "",
            ])


def append_xlsx(file_path: str, item: GroupItem) -> None:
    with csv_lock:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.append([
            item.index,
            item.name,
            item.url,
            item.subscribers if item.subscribers is not None else None,
            item.posting_status if item.posting_status is not None else None,
        ])
        wb.save(file_path)


def has_auth_cookie(page: Page) -> bool:
    try:
        cookies = page.context.cookies()
        for c in cookies:
            if c.get("name") == "remixsid" and c.get("value"):
                return True
    except Exception:
        pass
    try:
        cookie_str = page.evaluate("() => document.cookie")
        return ("remixsid=" in (cookie_str or ""))
    except Exception:
        return False


def looks_like_login_page(page: Page) -> bool:
    markers = [
        "Вход ВКонтакте",
        "QR-код",
        "Войти другим способом",
        "Создать аккаунт",
    ]
    try:
        content = page.content()
        return any(m in content for m in markers)
    except Exception:
        return False


def is_logged_in(page: Page) -> bool:
    if has_auth_cookie(page):
        return True
    if looks_like_login_page(page):
        return False
    try:
        page.wait_for_selector('a[href^="/feed"], a[href^="/im"], a[href^="/logout"]', timeout=1500)
        return True
    except PlaywrightTimeoutError:
        return False


def wait_for_login(page: Page, timeout_sec: int) -> None:
    start = time.time()
    while time.time() - start < timeout_sec:
        if is_logged_in(page):
            return
        time.sleep(1.0)
    raise TimeoutError("Авторизация не была завершена вовремя. Перезапустите и войдите в VK.")


def open_or_create_persistent_context(playwright: Playwright, user_data_dir: str, headless: bool) -> Tuple[BrowserContext, Page]:
    context = playwright.chromium.launch_persistent_context(
        user_data_dir=user_data_dir,
        headless=headless,
        viewport={"width": 1400, "height": 900},
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
        ],
    )
    page = context.pages[0] if context.pages else context.new_page()
    return context, page


def ensure_login_and_export_state(user_data_dir: str, headless: bool, login_timeout: int, state_file: str, auth_mode: str = "auto") -> None:
    """Prepare valid storage_state for subsequent new_context(storage_state=...).
    auth_mode: 'auto' | 'storage_only' | 'profile_only'
    - storage_only: use existing state_file if valid; do NOT open persistent profile or wait for login
    - profile_only: use persistent profile (.vk_user_data) to export state (may wait for login)
    - auto: try storage_state first; if invalid, try profile; if still invalid, wait for login (only if not headless)
    """
    from pathlib import Path
    state_path = Path(state_file)

    def validate_state_via_context(play: Playwright) -> bool:
        try:
            browser = play.chromium.launch(headless=True, args=["--no-sandbox", "--disable-blink-features=AutomationControlled"])  # always headless for check
            context = browser.new_context(storage_state=str(state_path), viewport={"width": 1400, "height": 900})
            page = context.new_page()
            try:
                page.goto("https://vk.com/")
                page.wait_for_timeout(800)
                ok = is_logged_in(page)
            finally:
                context.close()
                browser.close()
            return ok
        except Exception:
            return False

    # 1) storage_only: trust and skip
    if auth_mode == "storage_only":
        if not state_path.exists():
            raise RuntimeError("Требуется .vk_storage_state.json, но файл не найден. Перенесите его на сервер или используйте другой режим авторизации.")
        # Optional quick validation to fail fast if явно невалидно
        with sync_playwright() as p:
            if not validate_state_via_context(p):
                raise RuntimeError(".vk_storage_state.json невалиден/просрочен. Обновите файл авторизацией на локальной машине и скопируйте заново.")
        return

    # 2) auto/profile_only flows
    with sync_playwright() as p:
        # Try existing storage state first in auto-mode
        if auth_mode in ("auto",) and state_path.exists():
            if validate_state_via_context(p):
                return

        # If profile_only or storage invalid: try persistent profile and export
        context = p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=headless,
            viewport={"width": 1400, "height": 900},
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        try:
            page.goto("https://vk.com/")
        except Exception:
            pass

        if not is_logged_in(page):
            if auth_mode == "profile_only":
                # In profile_only we can wait (possibly under xvfb)
                print("Открылся VK. Пожалуйста, выполните вход (QR, пароль или через VK ID). Ожидаю завершения авторизации…", flush=True)
                wait_for_login(page, timeout_sec=login_timeout)
            else:
                # auto-mode and no login: if headless, do not block
                if headless:
                    context.close()
                    raise RuntimeError("Нет валидной авторизации (.vk_storage_state.json), а профиль без входа. Либо используйте --auth-mode storage_only с валидным state, либо выполните вход под xvfb.")
                print("Открылся VK. Пожалуйста, выполните вход (QR, пароль или через VK ID). Ожидаю завершения авторизации…", flush=True)
                wait_for_login(page, timeout_sec=login_timeout)

        # Export storage state
        context.storage_state(path=str(state_path))
        context.close()


def ensure_on_groups_page(page: Page) -> None:
    for _ in range(3):
        try:
            page.goto(VK_GROUPS_URL)
            try:
                page.wait_for_timeout(300)
                if not page.locator('input[data-testid="search_input"][placeholder="Поиск сообществ"]').count():
                    link = page.query_selector('a[href="/groups"], a[href*="/groups"]')
                    if link:
                        link.click()
                        page.wait_for_timeout(500)
            except Exception:
                pass
            loc = page.locator('input[data-testid="search_input"][placeholder="Поиск сообществ"]')
            loc.wait_for(state="visible", timeout=4000)
            return
        except Exception:
            continue
    try:
        page.goto(VK_GROUPS_URL)
    except Exception:
        pass


def navigate_to_communities_search(page: Page, query: str) -> None:
    ensure_on_groups_page(page)
    try:
        locator = page.locator('input[data-testid="search_input"][placeholder="Поиск сообществ"]')
        locator.wait_for(state="visible", timeout=5000)
        locator.click()
        locator.fill("")
        locator.type(query, delay=20)
        page.keyboard.press("Enter")
    except Exception:
        input_selectors = [
            'input[data-testid="search_input"]',
            '[role="searchbox"] input',
            'input[placeholder*="Поиск" i]',
            'input[type="search"]',
        ]
        for sel in input_selectors:
            try:
                el = page.query_selector(sel)
                if not el:
                    continue
                el.click()
                el.fill("")
                el.type(query, delay=20)
                page.keyboard.press("Enter")
                break
            except Exception:
                continue
        else:
            encoded = urllib.parse.quote(query)
            page.goto(SEARCH_URL_TEMPLATE.format(query=encoded))
    selectors = [
        'div[class*="search_results"]',
        'div[class*="SearchResults"]',
        'div[class*="_list"]',
        'div[class*="groups_list"]',
        'div:has(a.search_item__title)',
    ]
    for sel in selectors:
        try:
            page.wait_for_selector(sel, timeout=4000)
            break
        except PlaywrightTimeoutError:
            continue
    page.wait_for_timeout(800)


def parse_int_from_text(text: str) -> Optional[int]:
    if not text:
        return None
    t = text.lower().replace("\xa0", " ")
    m = re.search(r"(\d+[\s\d]*([\.,]\d+)?)\s*(тыс\.?|k)\b", t)
    if m:
        num = m.group(1).replace(" ", "").replace(",", ".")
        try:
            return int(float(num) * 1000)
        except Exception:
            pass
    m = re.search(r"(\d+[\s\d]*([\.,]\d+)?)\s*(млн\.?|m)\b", t)
    if m:
        num = m.group(1).replace(" ", "").replace(",", ".")
        try:
            return int(float(num) * 1000000)
        except Exception:
            pass
    numbers = re.findall(r"\d+[\s\d]*", t)
    if not numbers:
        return None
    def to_int(s: str) -> int:
        digits = re.sub(r"\D", "", s)
        return int(digits) if digits else 0
    values = [to_int(n) for n in numbers]
    return max(values) if values else None


def normalize_vk_url(url: str) -> str:
    if not url:
        return url
    try:
        parsed = urlparse(url)
        scheme = 'https'
        netloc = 'vk.com' if parsed.netloc.endswith('vk.ru') or parsed.netloc.endswith('vk.com') else parsed.netloc
        path = parsed.path or '/'
        # drop trailing slash except root
        if len(path) > 1 and path.endswith('/'):
            path = path[:-1]
        return urlunparse((scheme, netloc, path, '', '', ''))
    except Exception:
        return url


def extract_group_cards(page: Page) -> List[Tuple[str, str, Optional[int]]]:
    try:
        data = page.evaluate(
            """
            () => {
              function parseCount(text){
                if(!text) return null;
                const t = text.toLowerCase().replace(/\u00a0/g, ' ');
                let m = t.match(/(\d+[\s\d]*([\.,]\d+)?)\s*(тыс\.?|k)\b/);
                if(m){
                  const num = parseFloat(m[1].replace(/\s/g,'').replace(',','.'));
                  if(!isNaN(num)) return Math.round(num*1000);
                }
                m = t.match(/(\d+[\s\d]*([\.,]\d+)?)\s*(млн\.?|m)\b/);
                if(m){
                  const num = parseFloat(m[1].replace(/\s/g,'').replace(',', '.'));
                  if(!isNaN(num)) return Math.round(num*1000000);
                }
                const digits = (t.match(/\d+[\s\d]*/g)||[]).map(s=>parseInt(s.replace(/\D/g,''))).filter(n=>!isNaN(n));
                if(digits.length===0) return null;
                return Math.max(...digits);
              }
              function center(el){
                const r = el.getBoundingClientRect();
                return {x: r.left + r.width/2, y: r.top + r.height/2};
              }
              // Prefer anchors without query params by stripping location search/hash
              const anchors = Array.from(document.querySelectorAll("a[href^='/public'], a[href^='/club'], a[href^='/event'], a[href^='/community']"));
              const candidates = Array.from(document.querySelectorAll("span, div"))
                .filter(n => {
                  const t = (n.textContent||'').toLowerCase();
                  return t.includes('подпис') || t.includes('участ');
                });
              const spans = candidates.map(n => ({ node:n, c:center(n), text:(n.textContent||'').trim() }));
              const seen = new Set();
              const out = [];
              for(const a of anchors){
                let href = a.getAttribute('href')||'';
                const name = (a.textContent||a.getAttribute('aria-label')||'').trim();
                if(!href || !name) continue;
                // Strip search/hash
                try{
                  const u = new URL(href, 'https://vk.com');
                  href = u.pathname;
                }catch(e){ /* ignore */ }
                const full = href.startsWith('/') ? 'https://vk.com'+href : href;
                if(seen.has(full)) continue;
                seen.add(full);
                const ac = center(a);
                let best = null, bestD = Infinity;
                for(const s of spans){
                  const dy = Math.abs(s.c.y - ac.y);
                  const dx = Math.abs(s.c.x - ac.x);
                  const d = dy + dx*0.1;
                  if(dy < 80 && d < bestD){
                    best = s;
                    bestD = d;
                  }
                }
                let subs = null;
                if(best){
                  const val = parseCount(best.text);
                  if(typeof val === 'number' && val>0) subs = val;
                }
                out.push([name, full, subs]);
              }
              return out;
            }
            """
        )
        results: List[Tuple[str, str, Optional[int]]] = []
        for name, url, subs in data:
            results.append((str(name), normalize_vk_url(str(url)), int(subs) if subs is not None else None))
        return results
    except Exception:
        return []


def extract_posting_status(page: Page) -> Optional[str]:
    for _ in range(3):
        try:
            status = page.evaluate(
                """
                () => {
                  const getText = (el) => (el?.innerText || el?.textContent || "").trim();
                  const q = (sel) => Array.from(document.querySelectorAll(sel));
                  const createBtn = document.querySelector('button[data-testid="posting_create_post_button"]');
                  if (createBtn && getText(createBtn)) return getText(createBtn);
                  const candidates = [ ...q('button'), ...q('a'), ...q('[role="button"]') ];
                  for (const el of candidates) {
                    const t = getText(el).toLowerCase();
                    if (!t) continue;
                    if (t.includes('создать пост')) return getText(el);
                    if (t.includes('предложить') && (t.includes('новост') || t.includes('пост') || t.includes('запись'))) return getText(el);
                  }
                  const blocks = q('.PostingReactBlock__root, .PostingFormBlock__root, [data-testid^="posting_"]');
                  for (const b of blocks) {
                    const btn = b.querySelector('button, a, [role="button"]');
                    const t = getText(btn).toLowerCase();
                    if (t.includes('создать пост') || t.includes('предложить')) return getText(btn);
                  }
                  return null;
                }
                """
            )
            if status:
                return status
        except Exception:
            pass
        page.wait_for_timeout(700)
    # Fallback selectors
    selectors = [
        'button[data-testid="posting_create_post_button"]',
        'button:has-text("Создать пост")',
        'button:has-text("Предложить")',
        'a:has-text("Предложить новость")',
        '.PostingReactBlock__root button',
        '.PostingFormBlock__root button',
    ]
    for sel in selectors:
        try:
            el = page.query_selector(sel)
            if el:
                txt = (el.inner_text() or el.text_content() or "").strip()
                if txt:
                    return txt
        except Exception:
            continue
    return "Недоступно"


def infinite_scroll_and_collect(
    page: Page,
    already: Set[str],
    max_per_query: int,
    stop_event: threading.Event,
) -> List[Tuple[str, str, Optional[int]]]:
    collected_urls_local: Set[str] = set()
    results: List[Tuple[str, str, Optional[int]]] = []
    stagnant_rounds = 0
    unlimited = max_per_query is None or max_per_query <= 0

    while not stop_event.is_set() and (unlimited or len(results) < max_per_query) and stagnant_rounds < 20:
        if not unlimited and len(results) >= max_per_query:
            break
        found = extract_group_cards(page)
        new_count = 0
        for name, url, subs in found:
            if stop_event.is_set() or (not unlimited and len(results) >= max_per_query):
                break
            norm = normalize_vk_url(url)
            if norm in already or norm in collected_urls_local:
                continue
            collected_urls_local.add(norm)
            results.append((name, norm, subs))
            new_count += 1

        if new_count == 0:
            stagnant_rounds += 1
        else:
            stagnant_rounds = 0

        # Check for end-of-results marker
        try:
            end_marker = page.query_selector(".EmptyCell, .SearchExtraInfo, .DivorceSearchExtraInfo, [data-testid='search-empty-results']")
            if end_marker and end_marker.is_visible():
                break
        except Exception:
            pass

        try:
            page.evaluate("window.scrollBy(0, document.body.scrollHeight)")
            try:
                page.wait_for_load_state("networkidle", timeout=4000)
            except PlaywrightTimeoutError:
                page.wait_for_timeout(1000) # Fallback
        except Exception:
            pass
        page.wait_for_timeout(200) # Small delay to let UI render
    return results


def worker_discover_groups(
    query: str,
    state_file: str,
    headless: bool,
    max_per_query: int,
    existing_urls: Set[str],
    stop_event: threading.Event,
) -> List[Tuple[str, str, Optional[int]]]:
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=headless,
                args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
            )
            context = browser.new_context(storage_state=state_file, viewport={"width": 1400, "height": 900})
            page = context.new_page()
            try:
                navigate_to_communities_search(page, query)
                discovered = infinite_scroll_and_collect(
                    page=page,
                    already=existing_urls,
                    max_per_query=max_per_query,
                    stop_event=stop_event,
                )
                return discovered
            finally:
                context.close()
                browser.close()
    except Exception as e:
        if not stop_event.is_set():
            print(f"Поток поиска для '{query}' завершился с ошибкой: {e}", file=sys.stderr)
        return []


def run(
    query_list: List[str],
    output_csv: str,
    output_xlsx: str,
    user_data_dir: str,
    headless: bool,
    max_per_query: int,
    login_timeout: int,
    auth_mode: str,
) -> None:
    ensure_csv(output_csv)
    ensure_xlsx(output_xlsx)

    state_file = os.path.abspath(STORAGE_STATE_FILE)
    ensure_parent_dir(state_file)
    ensure_login_and_export_state(user_data_dir=user_data_dir, headless=headless, login_timeout=login_timeout, state_file=state_file, auth_mode=auth_mode)

    existing_urls: Set[str] = set()
    try:
        with open(output_csv, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                url = normalize_vk_url((row.get("Ссылка") or "").strip())
                if url:
                    existing_urls.add(url)
    except FileNotFoundError:
        pass

    current_index = 0
    try:
        with open(output_csv, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            # -1 because of header
            current_index = sum(1 for _ in reader) - 1
            if current_index < 0:
                current_index = 0
    except (FileNotFoundError, StopIteration):
        current_index = 0

    stop_event = threading.Event()

    # --- Phase 1: Discovery ---
    all_discovered: List[Tuple[str, str, Optional[int]]] = []
    print("Этап 1: Поиск сообществ по запросам...")
    try:
        with ThreadPoolExecutor(max_workers=max(1, len(query_list))) as executor:
            futures = {
                executor.submit(worker_discover_groups, query, state_file, headless, max_per_query, existing_urls, stop_event)
                for query in query_list
            }
            with tqdm(total=len(futures), desc="Поисковые запросы", unit="q") as pbar:
                for future in as_completed(futures):
                    try:
                        results_per_query = future.result()
                        if results_per_query:
                            all_discovered.extend(results_per_query)
                    except Exception as exc:
                        print(f"Ошибка в потоке поиска: {exc}", file=sys.stderr)
                    pbar.update(1)
    except KeyboardInterrupt:
        print("Остановка по Ctrl+C. Завершаю потоки поиска…", file=sys.stderr)
        stop_event.set()
        return

    # --- Deduplication & Writing ---
    unique_groups_to_write: List[Tuple[str, str, Optional[int]]] = []
    seen_urls_after_discover = set(existing_urls)
    for name, url, subs in all_discovered:
        if url not in seen_urls_after_discover:
            unique_groups_to_write.append((name, url, subs))
            seen_urls_after_discover.add(url)

    if stop_event.is_set():
        return

    if not unique_groups_to_write:
        print("Новых уникальных сообществ для записи не найдено.")
        return

    print(f"\nНайдено {len(unique_groups_to_write)} новых сообществ. Запись в файлы...")

    index_ref = [current_index]
    with tqdm(total=len(unique_groups_to_write), desc="Запись результатов", unit="grp") as pbar:
        for name, url, subs in unique_groups_to_write:
            with index_lock:
                index_ref[0] += 1
                idx = index_ref[0]
            item = GroupItem(index=idx, name=name, url=url, subscribers=subs, posting_status=None)
            append_csv(output_csv, item)
            append_xlsx(output_xlsx, item)
            pbar.update(1)


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Парсер сообществ VK: многопоточность по запросам, потоковое сохранение, подписчики и статус публикации.")
    parser.add_argument("--queries", nargs="*", default=[
        "Авто из Кореи",
        "Авто из Японии",
        "Авто из Китая",
    ], help="Список поисковых запросов. По умолчанию три предустановленных.")
    parser.add_argument("--csv", default="./output/groups.csv", help="Путь к CSV файлу для сохранения.")
    parser.add_argument("--xlsx", default="./output/groups.xlsx", help="Путь к Excel файлу для сохранения.")
    parser.add_argument("--user-data-dir", default="./.vk_user_data", help="Каталог для хранения сессии браузера (куки).")
    parser.add_argument("--headless", action="store_true", help="Запуск без интерфейса (для серверов/VDS используйте вместе с --auth-mode storage_only или xvfb-run).")
    parser.add_argument("--max-per-query", type=int, default=0, help="Максимум результатов на запрос. 0 или меньше = без лимита.")
    parser.add_argument("--login-timeout", type=int, default=300, help="Таймаут ожидания авторизации, сек.")
    parser.add_argument("--auth-mode", choices=["auto", "storage_only", "profile_only"], default="auto", help="Режим авторизации: auto/storage_only/profile_only.")

    args = parser.parse_args(list(argv) if argv is not None else None)

    csv_path = os.path.abspath(args.csv)
    xlsx_path = os.path.abspath(args.xlsx)
    user_data_dir = os.path.abspath(args.user_data_dir)

    try:
        run(
            query_list=args.queries,
            output_csv=csv_path,
            output_xlsx=xlsx_path,
            user_data_dir=user_data_dir,
            headless=args.headless,
            max_per_query=args.max_per_query,
            login_timeout=args.login_timeout,
            auth_mode=args.auth_mode,
        )
        print(f"Готово. CSV: {csv_path}, XLSX: {xlsx_path}")
        return 0
    except Exception as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
